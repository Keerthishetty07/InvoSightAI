import json
from io import BytesIO
from datetime import date
from decimal import Decimal
from xml.etree.ElementTree import Element, SubElement, tostring
from xml.dom import minidom

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def _money(value):
    return f"{Decimal(str(value or 0)):.2f}"


def _date(value):
    return value.strftime("%Y%m%d") if value else ""


def _invoice_dict(invoice):
    return {
        "Invoice Number": invoice.invoice_number or "",
        "Invoice Date": invoice.invoice_date.isoformat() if invoice.invoice_date else "",
        "Due Date": invoice.due_date.isoformat() if invoice.due_date else "",
        "Vendor": invoice.vendor or "",
        "Vendor GSTIN": invoice.vendor_gstin or "",
        "Customer": invoice.customer or "",
        "Customer GSTIN": invoice.customer_gstin or "",
        "Taxable Amount": _money(invoice.subtotal),
        "Tax Amount": _money(invoice.tax),
        "Invoice Total": _money(invoice.total),
        "Currency": invoice.currency or "INR",
        "Status": invoice.status or "",
    }


def export_tally_excel(invoice):
    """Create an Excel workbook suitable for TallyPrime user-defined mapping."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Transactions"

    headers = [
        "Voucher Date",
        "Voucher Type",
        "Voucher Number",
        "Party Ledger Name",
        "Party GSTIN",
        "Sales Ledger",
        "Taxable Amount",
        "Tax Amount",
        "Invoice Total",
        "Currency",
        "Due Date",
        "Narration",
    ]
    ws.append(headers)

    ws.append([
        invoice.invoice_date.isoformat() if invoice.invoice_date else "",
        "Sales",
        invoice.invoice_number or "",
        invoice.customer or invoice.vendor or "",
        invoice.customer_gstin or invoice.vendor_gstin or "",
        "Sales",
        float(invoice.subtotal or 0),
        float(invoice.tax or 0),
        float(invoice.total or 0),
        invoice.currency or "INR",
        invoice.due_date.isoformat() if invoice.due_date else "",
        f"InvoSightAI invoice {invoice.invoice_number or ''}".strip(),
    ])

    info = wb.create_sheet("README")
    notes = [
        ["InvoSightAI — TallyPrime Export"],
        ["Purpose", "Excel output for TallyPrime transaction import."],
        ["Import method", "TallyPrime > Alt+O > Import > Transactions."],
        ["Important", "This workbook uses TallyPrime field-oriented headers and can be mapped with a TallyPrime Mapping Template."],
        ["Party ledger", "Create/verify the party ledger in TallyPrime before importing."],
        ["Sales ledger", "The default exported ledger name is Sales; change the value if your company uses another ledger."],
        ["Line items", "The current InvoSightAI database stores invoice-level totals, not individual item rows. Item-level Tally vouchers should be added after line-item extraction is implemented."],
        ["Source", "Generated from the verified invoice record in InvoSightAI."],
    ]
    for row in notes:
        info.append(row)

    for sheet in (ws, info):
        for col in sheet.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            sheet.column_dimensions[get_column_letter(col[0].column)].width = min(max(max_len + 2, 12), 55)
        sheet.freeze_panes = "A2" if sheet is ws else None

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = (
        f'attachment; filename="InvoSightAI_Tally_{invoice.invoice_number or invoice.id}.xlsx"'
    )
    return response


def _ledger_entry(parent, ledger_name, amount, is_debit):
    entry = SubElement(parent, "ALLLEDGERENTRIES.LIST")
    SubElement(entry, "LEDGERNAME").text = ledger_name or ""
    SubElement(entry, "ISDEEMEDPOSITIVE").text = "Yes" if is_debit else "No"
    SubElement(entry, "AMOUNT").text = f"-{_money(amount)}" if is_debit else _money(amount)
    return entry


def export_tally_xml(invoice):
    """Create a TallyPrime XML voucher import document.

    This is intentionally invoice-level because the current Invoice model does not
    contain line-item/stock-item rows or separate CGST/SGST/IGST ledgers.
    """
    envelope = Element("ENVELOPE")
    header = SubElement(envelope, "HEADER")
    SubElement(header, "TALLYREQUEST").text = "Import Data"

    body = SubElement(envelope, "BODY")
    import_data = SubElement(body, "IMPORTDATA")
    request_desc = SubElement(import_data, "REQUESTDESC")
    SubElement(request_desc, "REPORTNAME").text = "Vouchers"
    static = SubElement(request_desc, "STATICVARIABLES")
    SubElement(static, "SVCURRENTCOMPANY").text = ""

    request_data = SubElement(import_data, "REQUESTDATA")
    message = SubElement(request_data, "TALLYMESSAGE")
    voucher = SubElement(
        message,
        "VOUCHER",
        {
            "VCHTYPE": "Sales",
            "ACTION": "Create",
            "OBJVIEW": "Invoice Voucher View",
        },
    )

    SubElement(voucher, "DATE").text = _date(invoice.invoice_date)
    SubElement(voucher, "VOUCHERTYPENAME").text = "Sales"
    SubElement(voucher, "VOUCHERNUMBER").text = invoice.invoice_number or ""
    SubElement(voucher, "REFERENCE").text = invoice.invoice_number or ""
    SubElement(voucher, "PARTYLEDGERNAME").text = invoice.customer or invoice.vendor or ""
    SubElement(voucher, "BASICBASEPARTYNAME").text = invoice.customer or invoice.vendor or ""
    SubElement(voucher, "NARRATION").text = (
        f"InvoSightAI invoice {invoice.invoice_number or ''}".strip()
    )

    # Party debit and a single Sales credit. Tax is included in the invoice total;
    # separate tax ledgers require CGST/SGST/IGST extraction and configuration.
    _ledger_entry(
        voucher,
        invoice.customer or invoice.vendor or "Party",
        invoice.total,
        True,
    )
    _ledger_entry(voucher, "Sales", invoice.total, False)

    xml_bytes = minidom.parseString(tostring(envelope, encoding="utf-8")).toprettyxml(
        indent="  ",
        encoding="utf-8",
    )

    response = HttpResponse(xml_bytes, content_type="application/xml")
    response["Content-Disposition"] = (
        f'attachment; filename="InvoSightAI_Tally_{invoice.invoice_number or invoice.id}.xml"'
    )
    return response


def export_tally_json(invoice):
    """Create a TallyPrime Release 7-style integration request payload.

    The payload keeps invoice fields explicit and is useful for Tally integration/
    transformation. Exact voucher fields can vary with the company's Tally setup.
    """
    payload = {
        "version": 1,
        "tallyrequest": "Import",
        "type": "Data",
        "id": "Vouchers",
        "static_variables": {
            "svVchImportFormat": "JSONEx",
        },
        "invoice": {
            "voucher_type": "Sales",
            "voucher_number": invoice.invoice_number or "",
            "date": invoice.invoice_date.isoformat() if invoice.invoice_date else "",
            "party_ledger_name": invoice.customer or invoice.vendor or "",
            "party_gstin": invoice.customer_gstin or invoice.vendor_gstin or "",
            "sales_ledger_name": "Sales",
            "taxable_amount": _money(invoice.subtotal),
            "tax_amount": _money(invoice.tax),
            "total_amount": _money(invoice.total),
            "currency": invoice.currency or "INR",
            "due_date": invoice.due_date.isoformat() if invoice.due_date else "",
        },
        "source": {
            "application": "InvoSightAI",
            "invoice_id": invoice.id,
        },
        "note": (
            "Current InvoSightAI stores invoice-level totals. "
            "Separate stock items and CGST/SGST/IGST ledger allocations require "
            "line-item and tax-component extraction."
        ),
    }

    response = HttpResponse(
        json.dumps(payload, indent=2, ensure_ascii=False),
        content_type="application/json; charset=utf-8",
    )
    response["Content-Disposition"] = (
        f'attachment; filename="InvoSightAI_Tally_{invoice.invoice_number or invoice.id}.json"'
    )
    return response
